import csv
import io
from datetime import datetime, timedelta, timezone
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func as sa_func

from app.auth.service import require_admin, require_executive, normalize_tracks
from app.database.connection import get_db
from app.database.models import (
    AttendancePoint, Employee, UserProgress, UserNote,
    TimeRecord, PerformanceReview, AbsenceRecord,
)
from app.content.loader import get_modules_for_tracks

router = APIRouter(prefix="/api/admin", tags=["admin"])

VALID_TRACKS = {"hr", "warehouse", "administrative", "management"}


# -- Schemas --

class EmployeeCreate(BaseModel):
    employee_id: str
    first_name: str
    last_name: str
    tracks: list[str]
    is_admin: bool = False
    department: str | None = None


class EmployeeUpdate(BaseModel):
    first_name: str | None = None
    last_name: str | None = None
    tracks: list[str] | None = None
    is_admin: bool | None = None
    is_manager: bool | None = None
    is_executive: bool | None = None
    manager_employee_id: str | None = None
    department: str | None = None


class EmployeeImportRow(BaseModel):
    employee_id: str
    track: str  # CSV column: single value or pipe-separated (e.g. "hr|warehouse")
    name: str | None = None
    full_name: str | None = None
    first_name: str | None = None
    last_name: str | None = None
    is_admin: bool = False
    department: str | None = None
    manager_employee_id: str | None = None
    location: str | None = None
    division: str | None = None


class EmployeeImportRequest(BaseModel):
    employees: list[EmployeeImportRow]


class AdminNoteUpdate(BaseModel):
    status: str | None = None  # "open" | "answered"
    admin_reply: str | None = None


# -- Helpers --

def _progress_summary(employee_id: str, db: Session) -> dict:
    rows = db.query(UserProgress).filter_by(employee_id=employee_id).all()
    completed = sum(1 for r in rows if r.module_completed)
    last_active = max((r.last_updated for r in rows if r.last_updated), default=None)
    return {
        "modules_completed": completed,
        "total_modules_seen": len(rows),
        "last_active": last_active.isoformat() if last_active else None,
    }


def _serialize(emp: Employee, db: Session) -> dict:
    return {
        "id": emp.id,
        "employee_id": emp.employee_id,
        "first_name": emp.first_name,
        "last_name": emp.last_name,
        "full_name": f"{emp.first_name} {emp.last_name}",
        "tracks": normalize_tracks(emp.track),
        "is_admin": emp.is_admin,
        "is_manager": bool(emp.is_manager),
        "is_executive": bool(emp.is_executive),
        "manager_employee_id": emp.manager_employee_id,
        "department": emp.department,
        "totp_enabled": bool(emp.totp_enabled),
        "created_at": emp.created_at.isoformat() if emp.created_at else None,
        "first_login_at": emp.first_login_at.isoformat() if emp.first_login_at else None,
        "last_login_at": emp.last_login_at.isoformat() if emp.last_login_at else None,
        "progress": _progress_summary(emp.employee_id, db),
    }


def _resolve_names(row: EmployeeImportRow) -> tuple[str | None, str | None]:
    if row.first_name and row.last_name:
        first_name = row.first_name.strip()
        last_name = row.last_name.strip()
        if first_name and last_name:
            return first_name, last_name

    raw_name = (row.full_name or row.name or "").strip()
    if not raw_name:
        return None, None

    parts = raw_name.split()
    if len(parts) < 2:
        return None, None

    return parts[0], " ".join(parts[1:])


# -- Routes --

@router.get("/employees")
def list_employees(
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    employees = db.query(Employee).order_by(Employee.created_at).all()
    return [_serialize(e, db) for e in employees]


@router.post("/employees", status_code=status.HTTP_201_CREATED)
def create_employee(
    payload: EmployeeCreate,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    tracks = normalize_tracks(payload.tracks)
    if not tracks:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tracks must include at least one of: {', '.join(sorted(VALID_TRACKS))}",
        )
    existing = db.query(Employee).filter_by(employee_id=payload.employee_id.strip()).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Employee ID '{payload.employee_id}' already exists.",
        )
    emp = Employee(
        employee_id=payload.employee_id.strip(),
        first_name=payload.first_name.strip(),
        last_name=payload.last_name.strip(),
        track=tracks,
        is_admin=payload.is_admin,
        department=payload.department.strip() if payload.department else None,
        created_at=datetime.now(timezone.utc),
    )
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return _serialize(emp, db)


@router.post("/employees/import-bamboo")
async def import_bamboo_employees(
    file: UploadFile = File(...),
    default_track: str = "warehouse",
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Import BambooHR Employee Division & Department xlsx.

    Creates missing employees (using *default_track*) and updates existing ones
    with location, division, department, and manager.

    Expected columns: Last name, First name | Employee # | Location |
                      Division | Department | Job Title | Reporting to
    """
    name = file.filename or ""
    if not name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be a .xlsx file.")

    raw_tracks = [t.strip().lower() for t in default_track.replace("|", ",").split(",")]
    tracks = normalize_tracks(raw_tracks)
    if not tracks:
        raise HTTPException(
            status_code=400,
            detail=f"default_track must be one of: {', '.join(sorted(VALID_TRACKS))}",
        )

    contents = await file.read()
    try:
        rows = _read_xlsx_admin(contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file.")

    if not rows:
        raise HTTPException(status_code=400, detail="No rows found in file.")

    LOCATION_ALIASES_BAMBOO: dict[str, str] = {
        "memphis, tn": "API Memphis",
    }

    # Build full-name → employee_id lookup for manager resolution
    all_emps = db.query(Employee).all()
    name_to_id: dict[str, str] = {
        f"{e.first_name} {e.last_name}".lower(): e.employee_id
        for e in all_emps
    }
    existing_by_id: dict[str, Employee] = {e.employee_id: e for e in all_emps}

    created_count = 0
    updated_count = 0
    skipped = 0
    manager_linked = 0
    errors: list[dict] = []

    for i, raw_row in enumerate(rows, start=2):
        # Normalise header keys for lookup
        row: dict[str, str] = {k.strip().lower(): str(v or "").strip() for k, v in raw_row.items()}

        # Employee number — column "employee #"
        emp_num_raw = row.get("employee #") or row.get("employee#") or ""
        # openpyxl may return floats like "12345.0" for numeric cells
        if emp_num_raw.endswith(".0"):
            emp_num_raw = emp_num_raw[:-2]
        emp_id = emp_num_raw.strip()
        if not emp_id:
            skipped += 1
            continue

        # Name — column header is literally "last name, first name"; value "Doe, Jane"
        name_raw = (
            row.get("last name, first name")
            or row.get("last name first name")
            or row.get("name")
            or ""
        )
        first_name: str | None = None
        last_name: str | None = None
        if name_raw:
            if "," in name_raw:
                parts = name_raw.split(",", 1)
                last_name = parts[0].strip()
                first_name = parts[1].strip()
            else:
                parts = name_raw.split()
                if len(parts) >= 2:
                    first_name = parts[0]
                    last_name = " ".join(parts[1:])

        location_raw = row.get("location", "")
        location: str | None = LOCATION_ALIASES_BAMBOO.get(location_raw.lower(), location_raw) or None
        division: str | None = row.get("division") or None
        department: str | None = row.get("department") or None
        reporting_to: str | None = row.get("reporting to") or None

        emp = existing_by_id.get(emp_id)

        if emp:
            # Update existing employee
            emp.location = location
            emp.division = division
            emp.department = department

            if reporting_to:
                mgr_id = name_to_id.get(reporting_to.lower())
                if mgr_id and mgr_id != emp.employee_id:
                    emp.manager_employee_id = mgr_id
                    manager_linked += 1

            updated_count += 1
        else:
            # Create new employee
            if not first_name or not last_name:
                errors.append({
                    "row": i,
                    "employee_id": emp_id,
                    "detail": "Cannot parse name — skipped.",
                })
                skipped += 1
                continue

            new_emp = Employee(
                employee_id=emp_id,
                first_name=first_name,
                last_name=last_name,
                track=tracks,
                is_admin=False,
                location=location,
                division=division,
                department=department,
                created_at=datetime.now(timezone.utc),
            )

            if reporting_to:
                mgr_id = name_to_id.get(reporting_to.lower())
                if mgr_id and mgr_id != emp_id:
                    new_emp.manager_employee_id = mgr_id
                    manager_linked += 1

            db.add(new_emp)
            # Keep in-memory lookups consistent for rest of file
            existing_by_id[emp_id] = new_emp
            name_to_id[f"{first_name} {last_name}".lower()] = emp_id
            created_count += 1

    db.commit()
    return {
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped,
        "manager_linked": manager_linked,
        "errors": errors[:20],
    }


@router.post("/employees/import")
def import_employees(
    payload: EmployeeImportRequest,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not payload.employees:
        raise HTTPException(status_code=400, detail="No employees were provided for import.")

    existing_ids = {employee_id for (employee_id,) in db.query(Employee.employee_id).all()}
    seen_ids: set[str] = set()
    created: list[Employee] = []
    errors: list[dict] = []

    for index, row in enumerate(payload.employees, start=1):
        employee_id = row.employee_id.strip()
        # Support pipe-separated tracks in CSV: "hr|warehouse"
        raw_tracks = [t.strip().lower() for t in row.track.replace("|", ",").split(",")]
        tracks = normalize_tracks(raw_tracks)

        if not employee_id:
            errors.append({"row": index, "employee_id": None, "detail": "Employee number is required."})
            continue

        if not tracks:
            errors.append({
                "row": index,
                "employee_id": employee_id,
                "detail": f"Track '{row.track}' is invalid. Must be one of: {', '.join(sorted(VALID_TRACKS))}.",
            })
            continue

        if employee_id in seen_ids:
            errors.append({
                "row": index,
                "employee_id": employee_id,
                "detail": "Employee number is duplicated in this import file.",
            })
            continue

        if employee_id in existing_ids:
            errors.append({
                "row": index,
                "employee_id": employee_id,
                "detail": "Employee already exists.",
            })
            continue

        first_name, last_name = _resolve_names(row)
        if not first_name or not last_name:
            errors.append({
                "row": index,
                "employee_id": employee_id,
                "detail": "A full employee name with first and last name is required.",
            })
            continue

        employee = Employee(
            employee_id=employee_id,
            first_name=first_name,
            last_name=last_name,
            track=tracks,
            is_admin=row.is_admin,
            department=row.department.strip() if row.department else None,
            manager_employee_id=row.manager_employee_id.strip() if row.manager_employee_id else None,
            location=row.location.strip() if row.location else None,
            division=row.division.strip() if row.division else None,
            created_at=datetime.now(timezone.utc),
        )
        db.add(employee)
        created.append(employee)
        seen_ids.add(employee_id)

    if created:
        db.commit()

    return {
        "added": len(created),
        "skipped": len(errors),
        "errors": errors,
    }


@router.post("/employees/import-xlsx")
async def import_employees_xlsx(
    file: UploadFile = File(...),
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Import employees from a standard xlsx using the same column layout as the CSV template.

    Expected columns: name, employee_id, track, is_admin (optional), department (optional),
                      manager_employee_id (optional), location (optional), division (optional)
    Also accepts a "Reporting to" column with manager full names — resolved to employee IDs
    after all rows are created so same-batch managers are found too.
    """
    fname = file.filename or ""
    if not fname.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be a .xlsx file.")

    contents = await file.read()
    try:
        raw_rows = _read_xlsx_admin(contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file.")

    if not raw_rows:
        raise HTTPException(status_code=400, detail="No rows found in file.")

    def _norm(h: str) -> str:
        return h.strip().lower().replace(" ", "_")

    employees: list[EmployeeImportRow] = []
    # emp_id -> lowercased manager full name, for second-pass resolution
    reporting_to_map: dict[str, str] = {}

    for row in raw_rows:
        r = {_norm(k): str(v or "").strip() for k, v in row.items()}
        name_val = r.get("name") or r.get("full_name") or ""
        emp_id = r.get("employee_id") or ""
        if emp_id.endswith(".0"):
            emp_id = emp_id[:-2]
        track = r.get("track") or ""
        is_admin_raw = r.get("is_admin", "false").lower()
        is_admin_val = is_admin_raw in ("true", "yes", "y", "1")
        dept = r.get("department") or r.get("dept") or ""
        loc = r.get("location") or ""
        division = r.get("division") or ""
        # Accept both ID-based and name-based manager columns
        mgr_id = r.get("manager_employee_id") or r.get("manager_id") or r.get("reports_to") or ""
        mgr_name = r.get("reporting_to") or ""  # name-based — resolved after import
        if mgr_name and emp_id:
            reporting_to_map[emp_id] = mgr_name.lower()
        employees.append(EmployeeImportRow(
            name=name_val or None,
            employee_id=emp_id,
            track=track,
            is_admin=is_admin_val,
            department=dept or None,
            manager_employee_id=mgr_id or None,
            location=loc or None,
            division=division or None,
        ))

    # First pass: create all employees
    result = import_employees(EmployeeImportRequest(employees=employees), admin, db)

    # Second pass: resolve "Reporting to" names now that all employees exist
    manager_linked = 0
    if reporting_to_map:
        all_emps = db.query(Employee).all()
        name_to_id: dict[str, str] = {
            f"{e.first_name} {e.last_name}".lower(): e.employee_id
            for e in all_emps
        }
        for emp_id_str, mgr_name_lower in reporting_to_map.items():
            mgr_id_resolved = name_to_id.get(mgr_name_lower)
            if not mgr_id_resolved:
                continue
            emp = db.query(Employee).filter_by(employee_id=emp_id_str).first()
            if emp and not emp.manager_employee_id:
                emp.manager_employee_id = mgr_id_resolved
                manager_linked += 1
        if manager_linked:
            db.commit()

    result["manager_linked"] = manager_linked
    return result


@router.patch("/employees/{employee_id}")
def update_employee(
    employee_id: str,
    payload: EmployeeUpdate,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    if payload.first_name is not None:
        emp.first_name = payload.first_name.strip()
    if payload.last_name is not None:
        emp.last_name = payload.last_name.strip()
    if payload.tracks is not None:
        tracks = normalize_tracks(payload.tracks)
        if not tracks:
            raise HTTPException(status_code=400, detail=f"Tracks must include at least one valid track.")
        emp.track = tracks
    if payload.is_admin is not None:
        emp.is_admin = payload.is_admin
    if payload.is_manager is not None:
        emp.is_manager = payload.is_manager
    if payload.is_executive is not None:
        emp.is_executive = payload.is_executive
    if payload.manager_employee_id is not None:
        emp.manager_employee_id = payload.manager_employee_id if payload.manager_employee_id != "" else None
    if payload.department is not None:
        emp.department = payload.department.strip() if payload.department.strip() else None
    db.commit()
    db.refresh(emp)
    return _serialize(emp, db)


@router.delete("/employees/{employee_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_employee(
    employee_id: str,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    # Prevent self-deletion
    if employee_id == admin.get("sub"):
        raise HTTPException(status_code=400, detail="You cannot delete your own account.")
    db.delete(emp)
    db.commit()


@router.post("/employees/{employee_id}/reset-progress", status_code=status.HTTP_200_OK)
def reset_employee_progress(
    employee_id: str,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete all progress records for an employee, resetting their journey."""
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    deleted = db.query(UserProgress).filter_by(employee_id=employee_id).delete()
    db.commit()
    return {"detail": f"Reset {deleted} progress records for {emp.first_name} {emp.last_name}."}


@router.post("/employees/{employee_id}/reset-totp", status_code=status.HTTP_200_OK)
def reset_employee_totp(
    employee_id: str,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Disable and clear TOTP for an employee so they can re-enroll."""
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    emp.totp_secret = None
    emp.totp_enabled = False
    db.commit()
    return {"detail": f"Two-factor authentication reset for {emp.first_name} {emp.last_name}."}


@router.get("/employees/{employee_id}/progress")
def get_employee_progress(
    employee_id: str,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Return detailed per-module progress for a specific employee."""
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    rows = db.query(UserProgress).filter_by(employee_id=employee_id).all()
    return [
        {
            "module_slug": r.module_slug,
            "visited": r.visited,
            "visited_at": r.visited_at.isoformat() if r.visited_at else None,
            "acknowledgements_completed": r.acknowledgements_completed,
            "quiz_passed": r.quiz_passed,
            "quiz_score": r.quiz_score,
            "quiz_attempts": r.quiz_attempts,
            "module_completed": r.module_completed,
            "completed_at": r.completed_at.isoformat() if r.completed_at else None,
        }
        for r in rows
    ]


@router.get("/employees/{employee_id}/notes")
def get_employee_notes(
    employee_id: str,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Return all notes written by a specific employee."""
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")
    notes = (
        db.query(UserNote)
        .filter_by(employee_id=employee_id)
        .order_by(UserNote.updated_at.desc())
        .all()
    )
    return [
        {
            "id": n.id,
            "module_slug": n.module_slug,
            "module_title": n.module_title,
            "note_text": n.note_text,
            "selected_text": n.selected_text,
            "anchor_id": n.anchor_id,
            "status": n.status,
            "admin_reply": n.admin_reply,
            "replied_by": n.replied_by,
            "replied_at": n.replied_at.isoformat() if n.replied_at else None,
            "created_at": n.created_at.isoformat() if n.created_at else None,
            "updated_at": n.updated_at.isoformat() if n.updated_at else None,
        }
        for n in notes
    ]


@router.patch("/employees/{employee_id}/notes/{note_id}")
def update_employee_note(
    employee_id: str,
    note_id: int,
    payload: AdminNoteUpdate,
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Allow admins to respond to or update status on an employee note."""
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")

    note = (
        db.query(UserNote)
        .filter(UserNote.id == note_id, UserNote.employee_id == employee_id)
        .first()
    )
    if not note:
        raise HTTPException(status_code=404, detail="Note not found.")

    if payload.status is not None:
        if payload.status not in {"open", "answered"}:
            raise HTTPException(status_code=422, detail="Status must be 'open' or 'answered'.")
        note.status = payload.status

    if payload.admin_reply is not None:
        cleaned_reply = payload.admin_reply.strip()
        note.admin_reply = cleaned_reply or None
        if note.admin_reply:
            note.replied_by = admin.get("sub")
            note.replied_at = datetime.now(timezone.utc)
            if payload.status is None:
                note.status = "answered"
        else:
            note.replied_by = None
            note.replied_at = None

    db.commit()
    db.refresh(note)

    return {
        "id": note.id,
        "module_slug": note.module_slug,
        "module_title": note.module_title,
        "note_text": note.note_text,
        "selected_text": note.selected_text,
        "anchor_id": note.anchor_id,
        "status": note.status,
        "admin_reply": note.admin_reply,
        "replied_by": note.replied_by,
        "replied_at": note.replied_at.isoformat() if note.replied_at else None,
        "created_at": note.created_at.isoformat() if note.created_at else None,
        "updated_at": note.updated_at.isoformat() if note.updated_at else None,
    }


# -- Attendance points import helpers (duplicated from manager router; no shared module) --

_POINTS_COL_MAP = {
    "employee #": "employee_id",
    "employee_number": "employee_id",
    "employee number": "employee_id",
    "employeenumber": "employee_id",
    "employee_id": "employee_id",
    "emp_id": "employee_id",
    "id": "employee_id",
    # Ignored lookup fields
    "last name": "_ignore",
    "last_name": "_ignore",
    "first name": "_ignore",
    "first_name": "_ignore",
    "location": "location",
    "point date": "point_date",
    "point_date": "point_date",
    "date": "point_date",
    "point": "point",
    "points": "point",
    "reason": "reason",
    "note": "note",
    "notes": "note",
    "flag code": "flag_code",
    "flag_code": "flag_code",
    "flagcode": "flag_code",
    "point total": "point_total",
    "point_total": "point_total",
    "total": "point_total",
    "running total": "point_total",
}


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace("-", "_")


def _parse_date(value: str) -> str | None:
    value = value.strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_float(value: str) -> float:
    try:
        return float(str(value).strip() or "0")
    except ValueError:
        return 0.0


def _map_row(raw_row: dict, col_map: dict) -> dict:
    mapped: dict = {}
    for raw_key, value in raw_row.items():
        canonical = col_map.get(_normalize_header(raw_key))
        if canonical:
            mapped[canonical] = value
    return mapped


def _read_csv_admin(contents: bytes) -> list[dict]:
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _read_xlsx_admin(contents: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows_data[0]]
    result = []
    for row in rows_data[1:]:
        if all(v is None for v in row):
            continue
        d: dict = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            if hasattr(val, "date") and callable(val.date):
                d[headers[i]] = val.date().isoformat()
            elif hasattr(val, "isoformat"):
                d[headers[i]] = val.isoformat()
            elif val is None:
                d[headers[i]] = ""
            else:
                d[headers[i]] = str(val).strip()
        result.append(d)
    return result


# -- Attendance points import --

@router.post("/import/points")
async def import_attendance_points(
    file: UploadFile = File(...),
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Upload weekly attendance point events (append-only historical ledger).
    Re-uploading the same file appends duplicate records — use DELETE to clear before re-uploading.

    Expected columns: Employee #, Location, Point Date, Point, Reason, Note, Flag Code, Point Total
    """
    name = file.filename or ""
    is_xlsx = name.lower().endswith(".xlsx")
    is_csv = name.lower().endswith(".csv")
    if not is_xlsx and not is_csv:
        raise HTTPException(status_code=400, detail="File must be a .xlsx or .csv file.")

    contents = await file.read()
    try:
        rows = _read_xlsx_admin(contents) if is_xlsx else _read_csv_admin(contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file.")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty.")

    # Collect valid employee IDs first so we can do a replace-per-employee
    seen_employees: set[str] = set()
    for raw_row in rows:
        row = _map_row(raw_row, _POINTS_COL_MAP)
        emp_id = str(row.get("employee_id", "")).strip()
        if emp_id:
            seen_employees.add(emp_id)

    # Delete existing records for every employee in the file before re-inserting
    for emp_id in seen_employees:
        db.query(AttendancePoint).filter_by(employee_id=emp_id).delete()
    db.flush()

    imported_at = datetime.now(timezone.utc)
    inserted = 0
    skipped = 0
    errors: list[dict] = []

    for i, raw_row in enumerate(rows, start=2):
        row = _map_row(raw_row, _POINTS_COL_MAP)

        employee_id = str(row.get("employee_id", "")).strip()
        if not employee_id:
            errors.append({"row": i, "detail": "Missing employee number / ID."})
            skipped += 1
            continue

        emp = db.query(Employee).filter_by(employee_id=employee_id).first()
        if not emp:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Employee '{employee_id}' not found."})
            skipped += 1
            continue

        point_date_raw = str(row.get("point_date", "")).strip()
        point_date = _parse_date(point_date_raw)
        if not point_date:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Cannot parse Point Date: '{point_date_raw}'."})
            skipped += 1
            continue

        flag_code_raw = str(row.get("flag_code", "")).strip()

        db.add(AttendancePoint(
            employee_id=employee_id,
            location=str(row.get("location", "")).strip() or None,
            point_date=point_date,
            point=_parse_float(str(row.get("point", "0"))),
            reason=str(row.get("reason", "")).strip() or None,
            note=str(row.get("note", "")).strip() or None,
            flag_code=flag_code_raw or None,
            point_total=_parse_float(str(row.get("point_total", "0"))),
            imported_at=imported_at,
        ))
        inserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped, "errors": errors}


@router.delete("/import/points", status_code=status.HTTP_200_OK)
def clear_attendance_points(
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete all attendance point records so HR can re-upload a corrected file."""
    deleted = db.query(AttendancePoint).delete()
    db.commit()
    return {"deleted": deleted}


# -- Employee directory import (BambooHR) --

def _parse_employee_directory(contents: bytes) -> list[dict]:
    """Parse BambooHR Employee Division and Department export.
    Expected columns: Last name First name, Employee #, Location, Division, Department, Job Title, Reporting to
    Normalizes known location aliases (e.g. 'Memphis, TN' -> 'API Memphis').
    """
    import openpyxl

    LOCATION_ALIASES: dict[str, str] = {
        "memphis, tn": "API Memphis",
    }

    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    def _col(row: tuple, name: str):
        try:
            return row[headers.index(name)]
        except ValueError:
            return None

    results = []
    for row in rows[1:]:
        emp_num = _col(row, "employee #")
        location = str(_col(row, "location") or "").strip()
        division = str(_col(row, "division") or "").strip()
        department = str(_col(row, "department") or "").strip()
        reporting_to = str(_col(row, "reporting to") or "").strip() or None

        if not emp_num:
            continue

        # Normalize known typos/aliases to canonical location names
        location = LOCATION_ALIASES.get(location.lower(), location)

        results.append({
            "employee_id": str(int(emp_num)) if isinstance(emp_num, (int, float)) else str(emp_num).strip(),
            "location":     location     or None,
            "division":     division     or None,
            "department":   department   or None,
            "reporting_to": reporting_to,
        })
    return results


@router.post("/import/employee-directory")
async def import_employee_directory(
    file: UploadFile = File(...),
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Upload BambooHR Employee Division & Department export to sync location,
    division, department, and manager on existing employees."""
    name = file.filename or ""
    if not name.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be a .xlsx file.")

    contents = await file.read()
    try:
        rows = _parse_employee_directory(contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file.")

    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows found in file.")

    # Build a full-name → employee_id lookup for manager resolution
    all_emps = db.query(Employee).all()
    name_to_id: dict[str, str] = {
        f"{e.first_name} {e.last_name}".lower(): e.employee_id
        for e in all_emps
    }

    updated = 0
    skipped = 0
    manager_linked = 0
    errors: list[dict] = []

    for i, row in enumerate(rows, start=2):
        emp = db.query(Employee).filter_by(employee_id=row["employee_id"]).first()
        if not emp:
            skipped += 1
            errors.append({"row": i, "employee_id": row["employee_id"], "detail": "Employee not found in system"})
            continue

        emp.location   = row["location"]
        emp.division   = row["division"]
        emp.department = row["department"]

        # Resolve manager by full name (case-insensitive)
        if row.get("reporting_to"):
            mgr_id = name_to_id.get(row["reporting_to"].lower())
            if mgr_id and mgr_id != emp.employee_id:
                emp.manager_employee_id = mgr_id
                manager_linked += 1

        updated += 1

    db.commit()
    return {
        "inserted":       updated,
        "skipped":        skipped,
        "manager_linked": manager_linked,
        "errors":         errors[:20],
    }


# -- Managers list --

@router.get("/managers")
def list_managers(
    admin: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    """Return all employees with is_manager=True, for the view-as-manager dropdown."""
    managers = (
        db.query(Employee)
        .filter_by(is_manager=True)
        .order_by(Employee.last_name, Employee.first_name)
        .all()
    )
    return [
        {
            "employee_id": m.employee_id,
            "full_name": f"{m.first_name} {m.last_name}",
            "department": m.department,
        }
        for m in managers
    ]


# -- Data clear endpoints --

@router.delete("/import/time", status_code=status.HTTP_200_OK)
def clear_time_records(
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete all uploaded time/hours records so HR can re-upload a corrected file."""
    deleted = db.query(TimeRecord).delete()
    db.commit()
    return {"deleted": deleted}


@router.delete("/import/reviews", status_code=status.HTTP_200_OK)
def clear_performance_reviews(
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete all uploaded performance review records so HR can re-upload a corrected file."""
    deleted = db.query(PerformanceReview).delete()
    db.commit()
    return {"deleted": deleted}


@router.delete("/import/absences", status_code=status.HTTP_200_OK)
def clear_absence_records(
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Delete all uploaded absence records so HR can re-upload a corrected file."""
    deleted = db.query(AbsenceRecord).delete()
    db.commit()
    return {"deleted": deleted}


# -- Dashboard --

@router.get("/dashboard")
def get_dashboard(
    admin: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Aggregated dashboard stats for HR admins."""
    employees = db.query(Employee).all()
    total = len(employees)

    # Count by track - employees with multiple tracks are counted once per track
    by_track: dict[str, int] = {}
    for emp in employees:
        for t in normalize_tracks(emp.track):
            by_track[t] = by_track.get(t, 0) + 1

    # Get all published non-management modules (journey modules)
    # Use HR track to get the full list including all shared modules
    all_modules = get_modules_for_tracks(["hr"])
    journey_modules = [m for m in all_modules if "management" not in m.get("tracks", [])]
    journey_slugs = {m["slug"] for m in journey_modules}
    total_journey_modules = len(journey_modules)

    # Get all non-management-only employees (journey participants)
    journey_employees = [e for e in employees if normalize_tracks(e.track) != ["management"]]
    journey_employee_ids = {e.employee_id for e in journey_employees}

    # Fetch all progress rows for journey modules
    all_progress = db.query(UserProgress).filter(
        UserProgress.module_slug.in_(journey_slugs)
    ).all() if journey_slugs else []

    # Build per-employee completion map
    emp_completed: dict[str, int] = {}  # employee_id -> completed count
    emp_visited: dict[str, bool] = {}   # employee_id -> has any progress
    for row in all_progress:
        if row.employee_id not in journey_employee_ids:
            continue
        if row.module_completed:
            emp_completed[row.employee_id] = emp_completed.get(row.employee_id, 0) + 1
        if row.visited:
            emp_visited[row.employee_id] = True

    all_complete = sum(1 for eid in journey_employee_ids if emp_completed.get(eid, 0) >= total_journey_modules) if total_journey_modules > 0 else 0
    in_progress = sum(1 for eid in journey_employee_ids if emp_visited.get(eid, False) and emp_completed.get(eid, 0) < total_journey_modules)
    not_started = len(journey_employee_ids) - len(emp_visited)

    # Recent logins (last 7 days)
    cutoff = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=7)
    recent_logins = []
    for emp in sorted(employees, key=lambda e: e.last_login_at or datetime.min, reverse=True):
        if emp.last_login_at and emp.last_login_at >= cutoff:
            recent_logins.append({
                "full_name": f"{emp.first_name} {emp.last_name}",
                "tracks": normalize_tracks(emp.track),
                "last_login_at": emp.last_login_at.isoformat(),
            })

    # Per-module completion counts (journey modules only)
    module_completion: dict[str, int] = {}
    for row in all_progress:
        if row.employee_id in journey_employee_ids and row.module_completed:
            module_completion[row.module_slug] = module_completion.get(row.module_slug, 0) + 1

    module_progress = []
    for m in journey_modules:
        module_progress.append({
            "module_slug": m["slug"],
            "title": m["title"],
            "completed": module_completion.get(m["slug"], 0),
            "total": len(journey_employee_ids),
        })

    return {
        "total_employees": total,
        "by_track": by_track,
        "completion": {
            "all_complete": all_complete,
            "in_progress": in_progress,
            "not_started": not_started,
        },
        "recent_logins": recent_logins,
        "module_progress": module_progress,
    }
