import io
import re
from datetime import datetime, date, timezone

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from app.auth.service import require_executive, require_hr_and_admin, normalize_tracks
from app.database.connection import get_db
from app.database.models import (
    ATTENDANCE_THRESHOLDS,
    AttendancePoint,
    Employee,
    TimeRecord,
    WoshReport,
)

router = APIRouter(prefix="/api/executive", tags=["executive"])


# ── WOSH Parsing ──────────────────────────────────────────────────────────────

def _find_sheet(wb, *fragments):
    """Return the first worksheet whose name contains any of the fragments (case-insensitive)."""
    for frag in fragments:
        for name in wb.sheetnames:
            if frag.lower() in name.lower():
                return wb[name]
    return None


def _parse_dashboard(ws):
    """Extract summary KPIs and top-employees table from the Dashboard sheet."""
    rows = list(ws.iter_rows(values_only=True))
    summary = {
        "total_violations": 0,
        "employees_affected": 0,
        "early_arrivals": 0,
        "late_departures": 0,
        "managers": 0,
        "generated_text": "",
    }

    # Row 2 (index 1): generated text with counts
    if len(rows) > 1 and rows[1][0]:
        text = str(rows[1][0])
        summary["generated_text"] = text
        m = re.search(r"(\d+)\s+managers?", text, re.IGNORECASE)
        if m:
            summary["managers"] = int(m.group(1))

    # Row 4 (index 3): KPI values at cols 0, 4, 8, 12
    if len(rows) > 3:
        r = rows[3]
        def _safe_int(val):
            try:
                return int(val)
            except (TypeError, ValueError):
                return 0
        summary["total_violations"]  = _safe_int(r[0]  if len(r) > 0  else None)
        summary["employees_affected"] = _safe_int(r[4]  if len(r) > 4  else None)
        summary["early_arrivals"]    = _safe_int(r[8]  if len(r) > 8  else None)
        summary["late_departures"]   = _safe_int(r[12] if len(r) > 12 else None)

    # Top employees — find the header row then read until None or footnote
    top_employees = []
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] == "Employee Name" and len(row) > 1 and row[1] == "Manager":
            header_idx = i
            break

    if header_idx is not None:
        for row in rows[header_idx + 1:]:
            if row[0] is None or (isinstance(row[0], str) and row[0].startswith("*")):
                break
            top_employees.append({
                "employee_name": str(row[0]),
                "manager":       str(row[1]) if row[1] else None,
                "total":         row[2],
                "early":         row[3],
                "late":          row[4],
            })

    return summary, top_employees


def _parse_chart_data(ws):
    """Parse _ChartData sheet: by_manager, by_type, by_day."""
    rows = list(ws.iter_rows(values_only=True))
    by_manager, by_type, by_day = [], [], []

    for row in rows[1:]:  # skip header
        if len(row) < 5:
            continue
        if row[0] is not None:
            by_manager.append({
                "manager":    str(row[0]),
                "early_only": int(row[1] or 0),
                "late_only":  int(row[2] or 0),
                "both":       int(row[3] or 0),
                "total":      int(row[4] or 0),
            })
        if len(row) > 7 and row[6] is not None:
            by_type.append({"type": str(row[6]), "count": int(row[7] or 0)})
        if len(row) > 10 and row[9] is not None:
            by_day.append({"day": str(row[9]), "count": int(row[10] or 0)})

    return by_manager, by_type, by_day


def _parse_by_manager(ws):
    """Parse 'By Manager' sheet into a list of manager objects with employee sub-rows."""
    result = []
    current = None

    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        cell0 = row[0]
        if cell0 is None:
            continue
        s = str(cell0).strip()

        # Manager header: only col A has a value and it contains "|"
        if row[1] is None and "|" in s:
            if current is not None:
                result.append(current)
            parts = [p.strip() for p in s.split("|")]
            name = parts[0]
            violations = 0
            emp_count = 0
            for p in parts[1:]:
                nums = [int(x) for x in p.split() if x.isdigit()]
                if "violation" in p.lower() and nums:
                    violations = nums[0]
                elif "employee" in p.lower() and nums:
                    emp_count = nums[0]
            current = {"manager": name, "violations": violations, "employee_count": emp_count, "rows": []}
            continue

        # Skip column headers and subtotal rows
        if s == "Employee Name" or "subtotal" in s.lower():
            continue

        # Employee data row
        if current is not None and row[1] is not None:
            current["rows"].append({
                "employee_name": str(row[0]).strip() if row[0] else None,
                "emp_num":       row[1],
                "department":    str(row[2]) if row[2] else None,
                "total":         row[3],
                "early_arrive":  row[4],
                "late_leave":    row[5],
                "extra_time":    str(row[6]) if row[6] is not None else None,
                "days_affected": str(row[7]) if row[7] is not None else None,
            })

    if current is not None:
        result.append(current)

    return result


def _parse_exceptions(ws):
    """Parse 'All Exceptions' sheet into row dicts; also returns week_start/end dates."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], None, None

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    exceptions = []
    date_objs = []

    for row in rows[1:]:
        if all(c is None for c in row):
            continue
        record = {}
        for h, val in zip(headers, row):
            if isinstance(val, datetime):
                record[h] = val.strftime("%Y-%m-%d")
                date_objs.append(val.date())
            else:
                record[h] = val
        exceptions.append(record)

    week_start = str(min(date_objs)) if date_objs else None
    week_end   = str(max(date_objs)) if date_objs else None

    return exceptions, week_start, week_end


def _parse_wosh_workbook(wb):
    """Parse a WOSH Excel workbook into structured data."""
    dashboard_ws = _find_sheet(wb, "dashboard")
    chart_ws     = _find_sheet(wb, "chartdata", "_chart")
    by_mgr_ws    = _find_sheet(wb, "by manager")
    exc_ws       = _find_sheet(wb, "all exception", "exception")

    summary, top_employees = _parse_dashboard(dashboard_ws) if dashboard_ws else ({}, [])
    by_manager_chart, by_type, by_day = _parse_chart_data(chart_ws) if chart_ws else ([], [], [])
    by_manager_detail = _parse_by_manager(by_mgr_ws) if by_mgr_ws else []
    exceptions, week_start, week_end = _parse_exceptions(exc_ws) if exc_ws else ([], None, None)

    return {
        "summary":           summary,
        "chart": {
            "by_manager": by_manager_chart,
            "by_type":    by_type,
            "by_day":     by_day,
        },
        "top_employees":     top_employees,
        "by_manager_detail": by_manager_detail,
        "exceptions":        exceptions,
    }, week_start, week_end


# ── Dashboard endpoint ────────────────────────────────────────────────────────

@router.get("/dashboard")
def executive_dashboard(
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    """Org-wide headcount and hours-by-department summary."""
    employees = db.query(Employee).all()

    total = len(employees)
    by_dept: dict[str, int] = {}
    by_track: dict[str, int] = {}
    managers = 0
    admins = 0
    executives = 0

    for emp in employees:
        dept = emp.department or "Unassigned"
        by_dept[dept] = by_dept.get(dept, 0) + 1
        for t in normalize_tracks(emp.track):
            by_track[t] = by_track.get(t, 0) + 1
        if emp.is_manager:   managers += 1
        if emp.is_admin:     admins += 1
        if emp.is_executive: executives += 1

    # Hours by department
    time_rows = (
        db.query(
            Employee.department,
            sa_func.sum(TimeRecord.regular_hours).label("regular"),
            sa_func.sum(TimeRecord.ot_hours).label("ot"),
            sa_func.sum(TimeRecord.vacation_hours).label("vacation"),
            sa_func.sum(TimeRecord.personal_hours).label("personal"),
            sa_func.sum(TimeRecord.other_hours).label("other"),
            sa_func.sum(TimeRecord.absent_w_point_hours).label("absent_w_point"),
            sa_func.sum(TimeRecord.protected_hours).label("protected"),
            sa_func.count(sa_func.distinct(TimeRecord.employee_id)).label("emp_count"),
        )
        .join(TimeRecord, Employee.employee_id == TimeRecord.employee_id)
        .group_by(Employee.department)
        .all()
    )

    hours_by_dept = [
        {
            "department":         row.department or "Unassigned",
            "employee_count":     row.emp_count,
            "regular_hours":      round(row.regular or 0, 2),
            "ot_hours":           round(row.ot or 0, 2),
            "vacation_hours":     round(row.vacation or 0, 2),
            "personal_hours":     round(row.personal or 0, 2),
            "other_hours":        round(row.other or 0, 2),
            "absent_w_point_hours": round(row.absent_w_point or 0, 2),
            "protected_hours":    round(row.protected or 0, 2),
        }
        for row in time_rows
    ]

    date_range_row = db.query(
        sa_func.min(TimeRecord.week_start).label("earliest"),
        sa_func.max(TimeRecord.week_start).label("latest"),
        sa_func.max(TimeRecord.imported_at).label("imported"),
    ).first()

    hours_date_range = None
    last_updated = None
    if date_range_row and date_range_row.earliest:
        hours_date_range = f"{date_range_row.earliest} – {date_range_row.latest}"
    if date_range_row and date_range_row.imported:
        last_updated = date_range_row.imported.isoformat()

    # Attendance thresholds
    subq = (
        db.query(
            AttendancePoint.employee_id,
            sa_func.max(AttendancePoint.point_total).label("max_total"),
        )
        .group_by(AttendancePoint.employee_id)
        .subquery()
    )
    threshold_counts = {"verbal": 0, "written": 0, "final": 0, "termination": 0}
    for row in db.query(subq).all():
        t = _classify_threshold(row.max_total)
        if t:
            threshold_counts[t] += 1

    # Company-wide totals for KPI row
    totals_row = db.query(
        sa_func.sum(TimeRecord.regular_hours).label("total_reg"),
        sa_func.sum(TimeRecord.ot_hours).label("total_ot"),
    ).first()

    return {
        "headcount": {
            "total":         total,
            "by_department": [{"department": k, "count": v} for k, v in sorted(by_dept.items())],
            "by_track":      by_track,
            "managers":      managers,
            "admins":        admins,
            "executives":    executives,
        },
        "hours_by_department": sorted(hours_by_dept, key=lambda x: x["department"]),
        "hours_date_range":    hours_date_range,
        "last_updated_hours":  last_updated,
        "attendance_thresholds": threshold_counts,
        "totals": {
            "regular_hours": round(totals_row.total_reg or 0, 2) if totals_row else 0,
            "ot_hours":      round(totals_row.total_ot  or 0, 2) if totals_row else 0,
        },
    }


@router.get("/hours-by-location")
def hours_by_location(
    week_start: str | None = Query(None),
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    """Hours (regular + OT) grouped by display location and department."""
    q = (
        db.query(
            Employee.location,
            Employee.division,
            Employee.department,
            sa_func.sum(TimeRecord.regular_hours).label("regular"),
            sa_func.sum(TimeRecord.ot_hours).label("ot"),
        )
        .join(TimeRecord, Employee.employee_id == TimeRecord.employee_id)
    )
    if week_start:
        q = q.filter(TimeRecord.week_start == week_start)
    rows = q.group_by(Employee.location, Employee.division, Employee.department).all()

    # Normalize location into 3 display buckets
    def _display_location(loc: str | None, div: str | None) -> str:
        if not loc or loc == "Remote":
            return "AAP"
        if loc == "AAP Scottsboro":
            return "AAP"
        if loc == "API Scottsboro":
            return "API Scottsboro"
        if loc == "API Memphis":
            return "API Memphis"
        return loc  # fallback

    # Accumulate: display_location → department → {regular, ot}
    groups: dict[str, dict[str, dict]] = {}
    for row in rows:
        display = _display_location(row.location, row.division)
        dept = row.department or "Unassigned"
        if display not in groups:
            groups[display] = {}
        if dept not in groups[display]:
            groups[display][dept] = {"regular_hours": 0.0, "ot_hours": 0.0}
        groups[display][dept]["regular_hours"] += row.regular or 0
        groups[display][dept]["ot_hours"] += row.ot or 0

    ORDER = ["AAP", "API Scottsboro", "API Memphis"]

    result = []
    for loc in ORDER:
        if loc not in groups:
            continue
        depts = [
            {"department": dept, "regular_hours": round(v["regular_hours"], 2), "ot_hours": round(v["ot_hours"], 2)}
            for dept, v in sorted(groups[loc].items())
        ]
        total_reg = round(sum(d["regular_hours"] for d in depts), 2)
        total_ot  = round(sum(d["ot_hours"]      for d in depts), 2)
        result.append({
            "location":      loc,
            "regular_hours": total_reg,
            "ot_hours":      total_ot,
            "departments":   depts,
        })

    # Any locations not in ORDER list
    for loc, dept_map in groups.items():
        if loc in ORDER:
            continue
        depts = [
            {"department": dept, "regular_hours": round(v["regular_hours"], 2), "ot_hours": round(v["ot_hours"], 2)}
            for dept, v in sorted(dept_map.items())
        ]
        result.append({
            "location":      loc,
            "regular_hours": round(sum(d["regular_hours"] for d in depts), 2),
            "ot_hours":      round(sum(d["ot_hours"]      for d in depts), 2),
            "departments":   depts,
        })

    return {"locations": result}


# ── Location normalizer (shared by headcount / pto / adherence) ───────────────

def _normalize_location(loc: str | None, div: str | None) -> str:
    """Map raw Employee.location to one of the 3 canonical display buckets."""
    if not loc or loc == "Remote":
        return "AAP"
    if loc == "AAP Scottsboro":
        return "AAP"
    if loc == "API Scottsboro":
        return "API Scottsboro"
    if loc == "API Memphis":
        return "API Memphis"
    return loc


# ── Headcount by location ─────────────────────────────────────────────────────

@router.get("/headcount")
def headcount_by_location(
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    """Active employee headcount grouped by display location → department."""
    rows = (
        db.query(
            Employee.location,
            Employee.division,
            Employee.department,
            sa_func.count(Employee.id).label("count"),
        )
        .group_by(Employee.location, Employee.division, Employee.department)
        .all()
    )

    groups: dict[str, dict[str, int]] = {}
    grand_total = 0
    for row in rows:
        display = _normalize_location(row.location, row.division)
        dept = row.department or "Unassigned"
        if display not in groups:
            groups[display] = {}
        groups[display][dept] = groups[display].get(dept, 0) + (row.count or 0)
        grand_total += row.count or 0

    ORDER = ["AAP", "API Scottsboro", "API Memphis"]
    result = []
    for loc in ORDER:
        if loc not in groups:
            continue
        depts = [{"department": d, "count": c} for d, c in sorted(groups[loc].items())]
        result.append({"location": loc, "total": sum(d["count"] for d in depts), "departments": depts})
    for loc, dept_map in groups.items():
        if loc not in ORDER:
            depts = [{"department": d, "count": c} for d, c in sorted(dept_map.items())]
            result.append({"location": loc, "total": sum(d["count"] for d in depts), "departments": depts})

    return {"total": grand_total, "by_location": result}


# ── PTO analytics ─────────────────────────────────────────────────────────────

@router.get("/pto-analytics")
def pto_analytics(
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    """Vacation + personal hours by display location and department."""
    rows = (
        db.query(
            Employee.location,
            Employee.division,
            Employee.department,
            sa_func.sum(TimeRecord.vacation_hours).label("vacation"),
            sa_func.sum(TimeRecord.personal_hours).label("personal"),
            sa_func.sum(TimeRecord.protected_hours).label("protected"),
            sa_func.count(sa_func.distinct(TimeRecord.employee_id)).label("emp_count"),
        )
        .join(TimeRecord, Employee.employee_id == TimeRecord.employee_id)
        .group_by(Employee.location, Employee.division, Employee.department)
        .all()
    )

    groups: dict[str, dict[str, dict]] = {}
    for row in rows:
        display = _normalize_location(row.location, row.division)
        dept = row.department or "Unassigned"
        if display not in groups:
            groups[display] = {}
        if dept not in groups[display]:
            groups[display][dept] = {"vacation": 0.0, "personal": 0.0, "protected": 0.0, "emp_count": 0}
        groups[display][dept]["vacation"]  += row.vacation  or 0
        groups[display][dept]["personal"]  += row.personal  or 0
        groups[display][dept]["protected"] += row.protected or 0
        groups[display][dept]["emp_count"] += row.emp_count or 0

    ORDER = ["AAP", "API Scottsboro", "API Memphis"]

    def _build(loc: str, dept_map: dict) -> dict:
        depts = [
            {
                "department":      d,
                "vacation_hours":  round(v["vacation"],  2),
                "personal_hours":  round(v["personal"],  2),
                "protected_hours": round(v["protected"], 2),
                "total_pto":       round(v["vacation"] + v["personal"], 2),
                "employee_count":  v["emp_count"],
            }
            for d, v in sorted(dept_map.items())
        ]
        vac = round(sum(d["vacation_hours"]  for d in depts), 2)
        per = round(sum(d["personal_hours"]  for d in depts), 2)
        pro = round(sum(d["protected_hours"] for d in depts), 2)
        return {
            "location":        loc,
            "vacation_hours":  vac,
            "personal_hours":  per,
            "protected_hours": pro,
            "total_pto":       round(vac + per, 2),
            "departments":     depts,
        }

    result = [_build(loc, groups[loc]) for loc in ORDER if loc in groups]
    for loc, dept_map in groups.items():
        if loc not in ORDER:
            result.append(_build(loc, dept_map))

    return {"total_pto": round(sum(l["total_pto"] for l in result), 2), "locations": result}


# ── Shift adherence ───────────────────────────────────────────────────────────

@router.get("/shift-adherence")
def shift_adherence(
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    """Per-manager team adherence ranked by OT rate + unexcused absence rate."""
    managers = db.query(Employee).filter(Employee.is_manager == True).all()
    if not managers:
        return {"managers": [], "top_manager": None, "top_score": None}

    result = []
    for mgr in managers:
        time_data = (
            db.query(
                sa_func.sum(TimeRecord.regular_hours).label("regular"),
                sa_func.sum(TimeRecord.ot_hours).label("ot"),
                sa_func.sum(TimeRecord.absent_w_point_hours).label("absent_pts"),
                sa_func.count(sa_func.distinct(TimeRecord.employee_id)).label("team_size"),
            )
            .join(Employee, TimeRecord.employee_id == Employee.employee_id)
            .filter(Employee.manager_employee_id == mgr.employee_id)
            .first()
        )
        if not time_data or not time_data.team_size:
            continue

        regular = float(time_data.regular or 0)
        ot      = float(time_data.ot      or 0)
        absent  = float(time_data.absent_pts or 0)
        total   = regular + ot
        if total == 0:
            continue

        ot_rate     = ot / total
        absent_rate = min(absent / total, 1.0)
        score       = round((1 - ot_rate) * (1 - absent_rate) * 100, 1)

        result.append({
            "manager_id":           mgr.employee_id,
            "manager_name":         f"{mgr.first_name} {mgr.last_name}",
            "department":           mgr.department,
            "location":             _normalize_location(mgr.location, mgr.division),
            "team_size":            time_data.team_size,
            "regular_hours":        round(regular, 2),
            "ot_hours":             round(ot, 2),
            "ot_rate":              round(ot_rate * 100, 1),
            "absent_w_point_hours": round(absent, 2),
            "adherence_score":      score,
        })

    result.sort(key=lambda x: x["adherence_score"], reverse=True)
    return {
        "managers":    result,
        "top_manager": result[0]["manager_name"]    if result else None,
        "top_score":   result[0]["adherence_score"] if result else None,
    }


def _classify_threshold(total: float | None) -> str | None:
    if total is None:
        return None
    for min_pts, label in sorted(ATTENDANCE_THRESHOLDS.items(), reverse=True):
        if total >= min_pts:
            return label
    return None


# ── WOSH endpoints ────────────────────────────────────────────────────────────

def _serialize_report(r: WoshReport, include_exceptions: bool = True) -> dict:
    pd = r.parsed_data or {}
    if not include_exceptions and pd:
        pd = {k: v for k, v in pd.items() if k != "exceptions"}
    return {
        "id":          r.id,
        "week_label":  r.week_label,
        "week_start":  r.week_start,
        "week_end":    r.week_end,
        "uploaded_at": r.uploaded_at.isoformat() if r.uploaded_at else None,
        "parsed_data": pd,
    }


@router.post("/wosh/upload")
async def upload_wosh(
    file: UploadFile = File(...),
    week_label: str = "",
    user: dict = Depends(require_hr_and_admin),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel workbook (.xlsx or .xlsm).")

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse Excel file. Make sure it is a valid .xlsx workbook.")

    parsed_data, week_start, week_end = _parse_wosh_workbook(wb)

    # Auto-generate week_label from date range if not provided
    if not week_label.strip() and week_start and week_end:
        try:
            s = date.fromisoformat(week_start)
            e = date.fromisoformat(week_end)
            if s.month == e.month:
                week_label = f"Week of {s.strftime('%b %-d')}–{e.strftime('%-d, %Y')}"
            else:
                week_label = f"Week of {s.strftime('%b %-d')} – {e.strftime('%b %-d, %Y')}"
        except Exception:
            week_label = f"{week_start} to {week_end}"

    # Store week_start/end in summary too
    if parsed_data.get("summary") is not None:
        parsed_data["summary"]["week_start"] = week_start
        parsed_data["summary"]["week_end"]   = week_end

    report = WoshReport(
        week_label  = week_label.strip() or None,
        week_start  = week_start,
        week_end    = week_end,
        parsed_data = parsed_data,
        uploaded_by = user.get("sub"),
        uploaded_at = datetime.now(timezone.utc),
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    exc_count = len(parsed_data.get("exceptions") or [])
    mgr_count = len(parsed_data.get("by_manager_detail") or [])
    return {
        "id":           report.id,
        "week_label":   report.week_label,
        "week_start":   report.week_start,
        "week_end":     report.week_end,
        "uploaded_at":  report.uploaded_at.isoformat(),
        "exceptions":   exc_count,
        "managers":     mgr_count,
        "summary":      parsed_data.get("summary"),
    }


@router.delete("/wosh", status_code=status.HTTP_200_OK)
def clear_wosh_reports(
    user: dict = Depends(require_hr_and_admin),
    db: Session = Depends(get_db),
):
    """Delete all WOSH reports so HR can re-upload a fresh set."""
    deleted = db.query(WoshReport).delete()
    db.commit()
    return {"deleted": deleted}


@router.get("/wosh/latest")
def get_wosh_latest(
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    report = db.query(WoshReport).order_by(WoshReport.uploaded_at.desc()).first()
    if not report:
        return None
    return _serialize_report(report)


@router.get("/wosh/history")
def get_wosh_history(
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    reports = db.query(WoshReport).order_by(WoshReport.uploaded_at.desc()).all()
    return [_serialize_report(r, include_exceptions=False) for r in reports]


@router.get("/wosh/{report_id}")
def get_wosh_report(
    report_id: int,
    user: dict = Depends(require_executive),
    db: Session = Depends(get_db),
):
    report = db.query(WoshReport).filter_by(id=report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="WOSH report not found.")
    return _serialize_report(report)
