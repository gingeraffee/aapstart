import csv
import io
from datetime import date, datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session

from app.auth.service import require_admin, require_manager, normalize_tracks
from app.database.connection import get_db
from app.database.models import (
    ATTENDANCE_THRESHOLDS, AttendancePoint,
    Employee, TimeRecord, PerformanceReview, UserProgress, AbsenceRecord,
)

router = APIRouter(prefix="/api/manager", tags=["manager"])

# ── Column name aliases accepted from HRIS CSV exports ────────────────────────

_TIME_COL_MAP = {
    # Employee identifier
    "employee_id": "employee_id",
    "employeeid": "employee_id",
    "emp_id": "employee_id",
    "employee_number": "employee_id",
    "employeenumber": "employee_id",
    "employee #": "employee_id",
    "employee number": "employee_id",
    "id": "employee_id",
    # Employee name — accepted but ignored (looked up from DB)
    "employee_name": "_ignore",
    "name": "_ignore",
    "full_name": "_ignore",
    "employee name": "_ignore",
    # Period start date (optional — defaults to upload date)
    "week_start": "week_start",
    "weekstart": "week_start",
    "week_beginning": "week_start",
    "period_start": "week_start",
    "period start": "week_start",
    "start_date": "week_start",
    "start date": "week_start",
    "week": "week_start",
    "date": "week_start",
    "pay_period_start": "week_start",
    # Regular hours
    "regular_hours": "regular_hours",
    "regular hours": "regular_hours",
    "reg_hours": "regular_hours",
    "reg hours": "regular_hours",
    "reg": "regular_hours",
    "hours": "regular_hours",
    "hours_worked": "regular_hours",
    "regular": "regular_hours",
    # Overtime
    "ot_hours": "ot_hours",
    "ot hours": "ot_hours",
    "overtime_hours": "ot_hours",
    "overtime hours": "ot_hours",
    "overtime": "ot_hours",
    "ot": "ot_hours",
    # Vacation
    "vacation_hours": "vacation_hours",
    "vacation hours": "vacation_hours",
    "vacation": "vacation_hours",
    "vac_hours": "vacation_hours",
    "vac hours": "vacation_hours",
    "vac": "vacation_hours",
    # Personal
    "personal_hours": "personal_hours",
    "personal hours": "personal_hours",
    "personal": "personal_hours",
    "pers_hours": "personal_hours",
    "pers hours": "personal_hours",
    "pers": "personal_hours",
    # Other / misc
    "other_hours": "other_hours",
    "other hours": "other_hours",
    "other": "other_hours",
    "misc_hours": "other_hours",
    "misc hours": "other_hours",
    "misc": "other_hours",
    # Legacy PTO catch-all (still accepted from old files)
    "pto_hours": "pto_hours",
    "pto hours": "pto_hours",
    "pto": "pto_hours",
    "leave_hours": "pto_hours",
    "leave": "pto_hours",
}

_REVIEW_COL_MAP = {
    "employee_id": "employee_id",
    "employeeid": "employee_id",
    "emp_id": "employee_id",
    "employee_number": "employee_id",
    "employeenumber": "employee_id",
    "employee #": "employee_id",
    "id": "employee_id",
    "review_type": "review_type",
    "type": "review_type",
    "review type": "review_type",
    "evaluation_type": "review_type",
    "performance_review_type": "review_type",
    "due_date": "due_date",
    "due date": "due_date",
    "scheduled_date": "due_date",
    "scheduled date": "due_date",
    "review_date": "due_date",
    "review date": "due_date",
    "review date due": "due_date",
    "date": "due_date",
    "completed": "completed",
    "status": "completed",
    "complete": "completed",
    "done": "completed",
    "review completed (y/n)": "completed",
    "review completed": "completed",
    "completed_date": "completed_date",
    "completion_date": "completed_date",
    "date_completed": "completed_date",
    "actual_date": "completed_date",
    "review_completed_date": "completed_date",
    # BambooHR columns we don't need
    "last name, first name": "_ignore",
    "reports to": "_ignore",
    "department": "_ignore",
    "location": "_ignore",
}


# ── Absence category normalisation ───────────────────────────────────────────
# Maps raw HRIS category strings (lowercased) to canonical bucket names.
# Any value not in this map falls through to "other".
CATEGORY_MAP: dict[str, str] = {
    "vacation": "vacation",
    "personal": "personal",
    "absent (w/ point)": "absent_w_point",
    "absent w/ point": "absent_w_point",
    "absent_w_point": "absent_w_point",
    "absent w/point": "absent_w_point",
    "absent(w/point)": "absent_w_point",
    "protected": "protected",
    "protected leave": "protected",
    "fmla": "protected",
    "jury duty": "protected",
    "bereavement": "protected",
    "sick": "protected",
    "sick leave": "protected",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_header(value: str) -> str:
    return value.strip().lower().replace("-", "_")


def _parse_date(value: str) -> str | None:
    """Return ISO date string or None. Accepts YYYY-MM-DD and M/D/YYYY."""
    value = value.strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _parse_bool(value: str) -> bool:
    return value.strip().lower() in {"true", "yes", "y", "1", "completed", "done", "complete"}


def _parse_float(value: str) -> float:
    try:
        return float(value.strip() or "0")
    except ValueError:
        return 0.0


def _map_row(raw_row: dict, col_map: dict) -> dict:
    """Remap CSV row keys using col_map aliases."""
    mapped: dict = {}
    for raw_key, value in raw_row.items():
        canonical = col_map.get(_normalize_header(raw_key))
        if canonical:
            mapped[canonical] = value
    return mapped


_ABSENCE_COL_MAP = {
    # Employee identifier
    "employee_id": "employee_id",
    "employeeid": "employee_id",
    "employee_number": "employee_id",
    "employeenumber": "employee_id",
    "employee #": "employee_id",
    "employee number": "employee_id",
    "emp_id": "employee_id",
    "id": "employee_id",
    # Employee name (stored for reference)
    "name": "employee_name",
    "employee_name": "employee_name",
    "full_name": "employee_name",
    "employee name": "employee_name",
    # Absence category
    "category": "category",
    "type": "category",
    "absence_type": "category",
    "leave_type": "category",
    # First day of absence
    "from": "from_date",
    "from_date": "from_date",
    "start_date": "from_date",
    "start date": "from_date",
    "absence_start": "from_date",
    # Last day of absence
    "to": "to_date",
    "to_date": "to_date",
    "end_date": "to_date",
    "end date": "to_date",
    "absence_end": "to_date",
    # Date the request was submitted
    "requested": "requested_date",
    "requested_date": "requested_date",
    "request_date": "requested_date",
    "submitted": "requested_date",
    "submitted_date": "requested_date",
    "date_requested": "requested_date",
    # Hours taken
    "time_off": "time_off_hours",
    "time off": "time_off_hours",
    "hours": "time_off_hours",
    "hours_taken": "time_off_hours",
    "absence_hours": "time_off_hours",
    "duration": "time_off_hours",
}


def _read_csv(contents: bytes) -> list[dict]:
    text = contents.decode("utf-8-sig")  # strip BOM if present
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def _read_xlsx(contents: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    sheet_name = "Time Off Used" if "Time Off Used" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows_data[0]]
    result = []
    for row in rows_data[1:]:
        if all(v is None for v in row):
            continue
        d: dict = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            # openpyxl returns datetime/date objects for date cells — convert to ISO strings
            if hasattr(val, "date") and callable(val.date):
                d[headers[i]] = val.date().isoformat()
            elif hasattr(val, "isoformat"):
                d[headers[i]] = val.isoformat()
            elif val is None:
                d[headers[i]] = ""
            else:
                d[headers[i]] = str(val).strip()
        result.append(d)
    return result


def _read_xlsx_generic(contents: bytes) -> list[dict]:
    """Standard header-row XLSX reader — works for BambooHR reviews and similar reports."""
    import openpyxl
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    sheet_name = "Report Data" if "Report Data" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_data = list(ws.iter_rows(values_only=True))
    if not rows_data:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows_data[0]]
    result = []
    for row in rows_data[1:]:
        if all(v is None for v in row):
            continue
        d: dict = {}
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            if hasattr(val, "date") and callable(val.date):
                d[headers[i]] = val.date().isoformat()
            elif hasattr(val, "isoformat"):
                d[headers[i]] = val.isoformat()
            elif val is None:
                d[headers[i]] = ""
            else:
                d[headers[i]] = str(val).strip()
        result.append(d)
    return result


def _parse_period_totals_xlsx(contents: bytes) -> tuple[list[dict], str]:
    """
    Parse a Payclock 'Period Totals Report' XLSX into a flat list of dicts.

    Layout expected:
      Row 1 : report title + timestamp
      Row 4 : date range cell, e.g. "04/01/26 Wed - 05/23/26 Sat"
      Row 8 : column headers  (Employee Name | … | Id | … | Reg | OT 1 | Vac | Personal | Other)
      Row 9+: one data row per employee
      Last  : "Employee Count: N …" totals row (skipped)

    Returns (rows, period_start_iso).
    """
    import re, openpyxl
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    ws = wb.active
    rows_data = list(ws.iter_rows(values_only=True))

    # Extract period start — scan the first 10 rows for a date range string like
    # "05/17/26 Sun - 05/23/26 Sat". Different Payclock report layouts put this
    # in different rows (row 4 for AAP/Memphis, row 5 for Scottsboro), so we
    # search rather than hardcode the row index.
    period_start = date.today().isoformat()
    _date_re = re.compile(r"(\d{2}/\d{2}/\d{2})")
    found_date = False
    for row in rows_data[:10]:
        if found_date:
            break
        for cell in row:
            if cell and isinstance(cell, str):
                m = _date_re.match(cell.strip())
                if m:
                    try:
                        period_start = datetime.strptime(m.group(1), "%m/%d/%y").date().isoformat()
                        found_date = True
                        break
                    except ValueError:
                        pass

    # Find the header row dynamically (first column = "Employee Name")
    header_row_idx = None
    for i, row in enumerate(rows_data):
        if row[0] and str(row[0]).strip().lower() == "employee name":
            header_row_idx = i
            break

    if header_row_idx is None:
        return [], period_start

    result = []
    for row in rows_data[header_row_idx + 1:]:
        name = row[0]
        if not name:
            continue
        if str(name).strip().lower().startswith("employee count"):
            break
        emp_id = str(row[2]).strip() if row[2] is not None else ""
        if not emp_id or emp_id == "None":
            continue
        result.append({
            "employee_id": emp_id,
            "week_start": period_start,
            "regular_hours": str(row[4] or 0),
            "ot_hours":      str(row[5] or 0),
            "vacation_hours": str(row[6] or 0),
            "personal_hours": str(row[7] or 0),
            "other_hours":   str(row[8] or 0),
        })

    return result, period_start


# ── Import endpoints ──────────────────────────────────────────────────────────

@router.post("/import/time")
async def import_time(
    file: UploadFile = File(...),
    manager: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Upload a Payclock Period Totals Report (XLSX) or a weekly hours CSV.
    Replaces existing records for any (employee_id, week_start) pair in the file.

    XLSX: Payclock 'Period Totals Report' — period start date is auto-detected from the header.
    CSV columns (flexible — see _TIME_COL_MAP for accepted aliases):
      employee_id, week_start, regular_hours, ot_hours, vacation_hours, personal_hours, other_hours
    """
    name = file.filename or ""
    is_xlsx = name.lower().endswith(".xlsx")
    is_csv = name.lower().endswith(".csv")
    if not is_xlsx and not is_csv:
        raise HTTPException(status_code=400, detail="File must be a .xlsx or .csv file.")

    contents = await file.read()
    period_start_from_file: str | None = None

    try:
        if is_xlsx:
            rows, period_start_from_file = _parse_period_totals_xlsx(contents)
        else:
            rows = _read_csv(contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file.")

    if not rows:
        raise HTTPException(status_code=400, detail="CSV file is empty.")

    imported_at = datetime.now(timezone.utc)
    inserted = 0
    skipped = 0
    errors: list[dict] = []

    for i, raw_row in enumerate(rows, start=2):
        row = _map_row(raw_row, _TIME_COL_MAP)

        employee_id = row.get("employee_id", "").strip()
        week_start_raw = row.get("week_start", "").strip()

        if not employee_id:
            errors.append({"row": i, "detail": "Missing employee_id."})
            skipped += 1
            continue

        # Validate employee exists in the system
        emp = db.query(Employee).filter_by(employee_id=employee_id).first()
        if not emp:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Employee '{employee_id}' not found — check the ID matches exactly what's in the system."})
            skipped += 1
            continue

        # week_start is optional — defaults to today when absent
        if week_start_raw:
            week_start = _parse_date(week_start_raw)
            if not week_start:
                errors.append({"row": i, "employee_id": employee_id, "detail": f"Cannot parse date: '{week_start_raw}'. Use YYYY-MM-DD or M/D/YYYY."})
                skipped += 1
                continue
        else:
            week_start = date.today().isoformat()

        # Delete existing record for this employee + period, then insert fresh
        db.query(TimeRecord).filter_by(
            employee_id=employee_id, week_start=week_start
        ).delete()

        db.add(TimeRecord(
            employee_id=employee_id,
            week_start=week_start,
            regular_hours=_parse_float(row.get("regular_hours", "0")),
            ot_hours=_parse_float(row.get("ot_hours", "0")),
            vacation_hours=_parse_float(row.get("vacation_hours", "0")),
            personal_hours=_parse_float(row.get("personal_hours", "0")),
            other_hours=_parse_float(row.get("other_hours", "0")),
            # Legacy pto_hours: carry forward from old files, otherwise 0
            pto_hours=_parse_float(row.get("pto_hours", "0")),
            imported_at=imported_at,
        ))
        inserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped, "errors": errors}


@router.post("/import/reviews")
async def import_reviews(
    file: UploadFile = File(...),
    manager: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Upload a BambooHR Performance Reviews XLSX or a custom CSV.
    Upserts on (employee_id, review_type, due_date). Skips already-completed reviews.

    XLSX: BambooHR 'Performance Reviews Due' export — uses 'Report Data' sheet.
    CSV columns (flexible — see _REVIEW_COL_MAP for accepted aliases):
      employee_id, review_type, due_date
    """
    name = file.filename or ""
    is_xlsx = name.lower().endswith(".xlsx")
    is_csv = name.lower().endswith(".csv")
    if not is_xlsx and not is_csv:
        raise HTTPException(status_code=400, detail="File must be a .xlsx or .csv file.")

    contents = await file.read()
    try:
        rows = _read_xlsx_generic(contents) if is_xlsx else _read_csv(contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file.")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty.")

    imported_at = datetime.now(timezone.utc)
    inserted = 0
    skipped = 0
    errors: list[dict] = []

    for i, raw_row in enumerate(rows, start=2):
        row = _map_row(raw_row, _REVIEW_COL_MAP)

        employee_id = str(row.get("employee_id", "")).strip()
        review_type = row.get("review_type", "").strip()
        due_date_raw = row.get("due_date", "").strip()

        if not employee_id:
            errors.append({"row": i, "detail": "Missing employee_id."})
            skipped += 1
            continue

        # Skip reviews already marked as completed
        if _parse_bool(row.get("completed", "false")):
            skipped += 1
            continue

        # Validate employee exists in the system
        emp = db.query(Employee).filter_by(employee_id=employee_id).first()
        if not emp:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Employee '{employee_id}' not found — check the ID matches exactly what's in the system."})
            skipped += 1
            continue

        if not review_type:
            errors.append({"row": i, "employee_id": employee_id, "detail": "Missing review_type."})
            skipped += 1
            continue

        due_date = _parse_date(due_date_raw)
        if not due_date:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Cannot parse due_date: '{due_date_raw}'."})
            skipped += 1
            continue

        # Upsert: delete existing match, then insert
        db.query(PerformanceReview).filter_by(
            employee_id=employee_id,
            review_type=review_type,
            due_date=due_date,
        ).delete()

        db.add(PerformanceReview(
            employee_id=employee_id,
            review_type=review_type,
            due_date=due_date,
            completed=False,
            completed_date=None,
            imported_at=imported_at,
        ))
        inserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped, "errors": errors}


@router.post("/import/absences")
async def import_absences(
    file: UploadFile = File(...),
    manager: dict = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """
    Upload a time-off/absence report (XLSX or CSV).
    Replaces all existing absence records for every employee present in the file.

    Expected columns (flexible — see _ABSENCE_COL_MAP for accepted aliases):
      Employee Number, Name, Category, From, To, Requested, Time Off
    """
    name = file.filename or ""
    is_xlsx = name.lower().endswith(".xlsx")
    is_csv = name.lower().endswith(".csv")
    if not is_xlsx and not is_csv:
        raise HTTPException(status_code=400, detail="File must be a .xlsx or .csv file.")

    contents = await file.read()
    try:
        rows = _read_xlsx(contents) if is_xlsx else _read_csv(contents)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse file.")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty.")

    imported_at = datetime.now(timezone.utc)
    inserted = 0
    skipped = 0
    errors: list[dict] = []

    # Collect valid employee IDs first so we can do a bulk delete before insert
    seen_employee_ids: set[str] = set()
    parsed_rows: list[dict] = []

    for i, raw_row in enumerate(rows, start=2):
        row = _map_row(raw_row, _ABSENCE_COL_MAP)

        employee_id = str(row.get("employee_id", "")).strip()
        if not employee_id:
            errors.append({"row": i, "detail": "Missing employee number / ID."})
            skipped += 1
            continue

        emp = db.query(Employee).filter_by(employee_id=employee_id).first()
        if not emp:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Employee '{employee_id}' not found — check the ID matches exactly."})
            skipped += 1
            continue

        from_raw = str(row.get("from_date", "")).strip()
        requested_raw = str(row.get("requested_date", "")).strip()

        from_date = _parse_date(from_raw)
        if not from_date:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Cannot parse From date: '{from_raw}'."})
            skipped += 1
            continue

        requested_date = _parse_date(requested_raw)
        if not requested_date:
            errors.append({"row": i, "employee_id": employee_id, "detail": f"Cannot parse Requested date: '{requested_raw}'."})
            skipped += 1
            continue

        to_date = _parse_date(str(row.get("to_date", "")).strip())
        time_off_hours = _parse_float(str(row.get("time_off_hours", "0")))
        is_planned = requested_date < from_date

        seen_employee_ids.add(employee_id)
        parsed_rows.append({
            "employee_id": employee_id,
            "employee_name": str(row.get("employee_name", "")).strip() or None,
            "category": str(row.get("category", "")).strip() or None,
            "from_date": from_date,
            "to_date": to_date,
            "requested_date": requested_date,
            "time_off_hours": time_off_hours,
            "is_planned": is_planned,
        })

    # Replace absence records for seen employees
    if seen_employee_ids:
        db.query(AbsenceRecord).filter(
            AbsenceRecord.employee_id.in_(seen_employee_ids)
        ).delete(synchronize_session=False)

    for pr in parsed_rows:
        db.add(AbsenceRecord(
            employee_id=pr["employee_id"],
            employee_name=pr["employee_name"],
            category=pr["category"],
            from_date=pr["from_date"],
            to_date=pr["to_date"],
            requested_date=pr["requested_date"],
            time_off_hours=pr["time_off_hours"],
            is_planned=pr["is_planned"],
            imported_at=imported_at,
        ))
        inserted += 1

    # ── Aggregate hours into TimeRecord ──────────────────────────────────────
    # Zero out all absence-derived columns for seen employees — rebuilt below
    if seen_employee_ids:
        db.query(TimeRecord).filter(
            TimeRecord.employee_id.in_(seen_employee_ids)
        ).update(
            {
                "vacation_hours": 0.0,
                "personal_hours": 0.0,
                "other_hours": 0.0,
                "absent_w_point_hours": 0.0,
                "protected_hours": 0.0,
            },
            synchronize_session=False,
        )

    # Aggregate per (employee_id, week_start) using normalised category buckets
    from collections import defaultdict
    week_buckets: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {"vacation": 0.0, "personal": 0.0, "absent_w_point": 0.0, "protected": 0.0, "other": 0.0}
    )
    for pr in parsed_rows:
        from_date_obj = date.fromisoformat(pr["from_date"])
        week_start_obj = from_date_obj - timedelta(days=from_date_obj.weekday())
        week_start = week_start_obj.isoformat()
        norm_cat = CATEGORY_MAP.get((pr["category"] or "").strip().lower(), "other")
        week_buckets[(pr["employee_id"], week_start)][norm_cat] += pr["time_off_hours"]

    hours_upserted = 0
    for (employee_id, week_start), hrs in week_buckets.items():
        existing = db.query(TimeRecord).filter_by(
            employee_id=employee_id, week_start=week_start
        ).first()
        if existing:
            existing.vacation_hours       = round(hrs["vacation"], 2)
            existing.personal_hours       = round(hrs["personal"], 2)
            existing.other_hours          = round(hrs["other"], 2)
            existing.absent_w_point_hours = round(hrs["absent_w_point"], 2)
            existing.protected_hours      = round(hrs["protected"], 2)
        else:
            db.add(TimeRecord(
                employee_id=employee_id,
                week_start=week_start,
                regular_hours=0.0,
                ot_hours=0.0,
                vacation_hours=round(hrs["vacation"], 2),
                personal_hours=round(hrs["personal"], 2),
                other_hours=round(hrs["other"], 2),
                absent_w_point_hours=round(hrs["absent_w_point"], 2),
                protected_hours=round(hrs["protected"], 2),
                imported_at=imported_at,
            ))
        hours_upserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped, "errors": errors, "hours_upserted": hours_upserted}


# ── Dashboard helpers ─────────────────────────────────────────────────────────

def _fmt_iso_date(iso: str) -> str:
    """Format ISO date as "Mon D" cross-platform."""
    d = date.fromisoformat(iso)
    return d.strftime("%b") + " " + str(d.day)


def _month_bounds(month_str: str) -> tuple[str, str]:
    """Return (month_start, month_end) ISO strings for a 'YYYY-MM' input."""
    import calendar as _cal
    year, mon = int(month_str[:4]), int(month_str[5:7])
    last_day = _cal.monthrange(year, mon)[1]
    return f"{year}-{mon:02d}-01", f"{year}-{mon:02d}-{last_day:02d}"


def _threshold_for(point_total: float | None) -> str | None:
    if point_total is None:
        return None
    for pts, label in sorted(ATTENDANCE_THRESHOLDS.items(), reverse=True):
        if point_total >= pts:
            return label
    return None


def _build_snapshot(
    team: list,
    team_ids: set[str],
    team_by_id: dict,
    db: Session,
    month_start: str,
    month_end: str,
    pts_by_emp: dict[str, float],
    monday_flag_ids: set[str],
) -> dict:
    """Build a single dashboard snapshot for a given month range."""

    # ── Hours ────────────────────────────────────────────────────────────────
    time_records = db.query(TimeRecord).filter(
        TimeRecord.employee_id.in_(team_ids),
        TimeRecord.week_start >= month_start,
        TimeRecord.week_start <= month_end,
    ).all()

    hours_by_emp: dict[str, dict] = {
        eid: {
            "employee_id": eid,
            "full_name": f"{team_by_id[eid].first_name} {team_by_id[eid].last_name}",
            "department": team_by_id[eid].department,
            "regular_hours": 0.0,
            "ot_hours": 0.0,
            "vacation_hours": 0.0,
            "personal_hours": 0.0,
            "other_hours": 0.0,
            "absent_w_point_hours": 0.0,
            "protected_hours": 0.0,
            "weeks_included": 0,
        }
        for eid in team_ids
    }

    all_week_starts: set[str] = set()
    for r in time_records:
        if r.employee_id not in hours_by_emp:
            continue
        h = hours_by_emp[r.employee_id]
        h["regular_hours"] += r.regular_hours
        h["ot_hours"] += r.ot_hours
        vac  = getattr(r, "vacation_hours", 0.0) or 0.0
        pers = getattr(r, "personal_hours", 0.0) or 0.0
        other = getattr(r, "other_hours", 0.0) or 0.0
        awp  = getattr(r, "absent_w_point_hours", 0.0) or 0.0
        prot = getattr(r, "protected_hours", 0.0) or 0.0
        pto  = getattr(r, "pto_hours", 0.0) or 0.0
        # Legacy fallback: if all new cols are 0, put pto_hours into vacation
        if vac == 0.0 and pers == 0.0 and other == 0.0 and awp == 0.0 and prot == 0.0 and pto > 0.0:
            vac = pto
        h["vacation_hours"]       += vac
        h["personal_hours"]       += pers
        h["other_hours"]          += other
        h["absent_w_point_hours"] += awp
        h["protected_hours"]      += prot
        h["weeks_included"]       += 1
        all_week_starts.add(r.week_start)

    for h in hours_by_emp.values():
        for key in ("regular_hours", "ot_hours", "vacation_hours", "personal_hours",
                    "other_hours", "absent_w_point_hours", "protected_hours"):
            h[key] = round(h[key], 1)

    sorted_weeks = sorted(all_week_starts)
    if sorted_weeks:
        hours_date_range = f"{_fmt_iso_date(sorted_weeks[0])} – {_fmt_iso_date(sorted_weeks[-1])}"
        hours_week_count = len(sorted_weeks)
    else:
        hours_date_range = None
        hours_week_count = 0

    all_time = db.query(TimeRecord).filter(TimeRecord.employee_id.in_(team_ids)).all()
    last_updated_time = max((r.imported_at for r in all_time if r.imported_at), default=None)

    # ── Performance reviews (all incomplete, not month-scoped) ───────────────
    reviews = db.query(PerformanceReview).filter(
        PerformanceReview.employee_id.in_(team_ids),
    ).all()

    today = date.today()
    upcoming: list[dict] = []
    past_due: list[dict] = []

    for r in reviews:
        if r.completed:
            continue
        emp = team_by_id.get(r.employee_id)
        if not emp:
            continue
        try:
            due = date.fromisoformat(r.due_date)
        except ValueError:
            continue
        diff = (due - today).days
        entry = {
            "employee_id": r.employee_id,
            "full_name": f"{emp.first_name} {emp.last_name}",
            "review_type": r.review_type,
            "due_date": r.due_date,
        }
        if diff >= 0:
            upcoming.append({**entry, "days_until": diff})
        else:
            past_due.append({**entry, "days_overdue": abs(diff)})

    upcoming.sort(key=lambda x: x["days_until"])
    past_due.sort(key=lambda x: x["days_overdue"], reverse=True)
    last_updated_reviews = max((r.imported_at for r in reviews if r.imported_at), default=None)

    # ── Absences (month-scoped) ───────────────────────────────────────────────
    absence_records = db.query(AbsenceRecord).filter(
        AbsenceRecord.employee_id.in_(team_ids),
        AbsenceRecord.from_date >= month_start,
        AbsenceRecord.from_date <= month_end,
    ).all()

    absence_by_emp: dict[str, dict] = {}
    absence_category_by_emp: dict[str, dict] = {}
    all_absence_dates: list[str] = []

    for r in absence_records:
        emp = team_by_id.get(r.employee_id)
        if not emp:
            continue
        full_name = f"{emp.first_name} {emp.last_name}"
        norm_cat = CATEGORY_MAP.get((r.category or "").strip().lower(), "other")

        if r.employee_id not in absence_by_emp:
            absence_by_emp[r.employee_id] = {
                "employee_id": r.employee_id,
                "full_name": full_name,
                "planned_count": 0,
                "unplanned_count": 0,
                "planned_hours": 0.0,
                "unplanned_hours": 0.0,
            }
        if r.is_planned:
            absence_by_emp[r.employee_id]["planned_count"] += 1
            absence_by_emp[r.employee_id]["planned_hours"] += r.time_off_hours
        else:
            absence_by_emp[r.employee_id]["unplanned_count"] += 1
            absence_by_emp[r.employee_id]["unplanned_hours"] += r.time_off_hours

        if r.employee_id not in absence_category_by_emp:
            absence_category_by_emp[r.employee_id] = {
                "employee_id": r.employee_id,
                "full_name": full_name,
                "vacation_hours": 0.0,
                "personal_hours": 0.0,
                "absent_w_point_hours": 0.0,
                "protected_hours": 0.0,
                "other_hours": 0.0,
                "planned_count": 0,
                "unplanned_count": 0,
                "planned_hours": 0.0,
                "unplanned_hours": 0.0,
                "has_monday_flag": r.employee_id in monday_flag_ids,
            }
        cat_entry = absence_category_by_emp[r.employee_id]
        bucket = norm_cat + "_hours" if norm_cat != "other" else "other_hours"
        if bucket not in cat_entry:
            bucket = "other_hours"
        cat_entry[bucket] = round(cat_entry[bucket] + r.time_off_hours, 1)
        if r.is_planned:
            cat_entry["planned_count"] += 1
            cat_entry["planned_hours"] = round(cat_entry["planned_hours"] + r.time_off_hours, 1)
        else:
            cat_entry["unplanned_count"] += 1
            cat_entry["unplanned_hours"] = round(cat_entry["unplanned_hours"] + r.time_off_hours, 1)

        all_absence_dates.append(r.from_date)

    for d in absence_by_emp.values():
        d["planned_hours"] = round(d["planned_hours"], 1)
        d["unplanned_hours"] = round(d["unplanned_hours"], 1)

    sorted_abs = sorted(all_absence_dates)
    if sorted_abs:
        absence_date_range = f"{_fmt_iso_date(sorted_abs[0])} – {_fmt_iso_date(sorted_abs[-1])}"
    else:
        absence_date_range = None

    last_updated_absences = max(
        (r.imported_at for r in absence_records if r.imported_at), default=None
    )

    # ── Team roster ───────────────────────────────────────────────────────────
    progress_rows = db.query(UserProgress).filter(
        UserProgress.employee_id.in_(team_ids)
    ).all()
    emp_modules_completed: dict[str, int] = {}
    for row in progress_rows:
        if row.module_completed:
            emp_modules_completed[row.employee_id] = emp_modules_completed.get(row.employee_id, 0) + 1

    # Indirect reports: fetch direct reports of any manager on this team
    manager_ids_on_team = {e.employee_id for e in team if e.is_manager}
    sub_reports_by_mgr: dict[str, list] = {mid: [] for mid in manager_ids_on_team}
    if manager_ids_on_team:
        sub_emps = (
            db.query(Employee)
            .filter(
                Employee.manager_employee_id.in_(manager_ids_on_team),
                Employee.terminated_date.is_(None),
            )
            .order_by(Employee.last_name, Employee.first_name)
            .all()
        )
        sub_ids = {e.employee_id for e in sub_emps}
        sub_progress = db.query(UserProgress).filter(
            UserProgress.employee_id.in_(sub_ids)
        ).all() if sub_ids else []
        sub_modules: dict[str, int] = {}
        for row in sub_progress:
            if row.module_completed:
                sub_modules[row.employee_id] = sub_modules.get(row.employee_id, 0) + 1
        for e in sub_emps:
            if e.manager_employee_id in sub_reports_by_mgr:
                sub_reports_by_mgr[e.manager_employee_id].append({
                    "employee_id": e.employee_id,
                    "full_name": f"{e.first_name} {e.last_name}",
                    "tracks": normalize_tracks(e.track),
                    "department": e.department,
                    "last_login_at": e.last_login_at.isoformat() if e.last_login_at else None,
                    "first_login_at": e.first_login_at.isoformat() if e.first_login_at else None,
                    "modules_completed": sub_modules.get(e.employee_id, 0),
                })

    team_list = [
        {
            "employee_id": e.employee_id,
            "full_name": f"{e.first_name} {e.last_name}",
            "tracks": normalize_tracks(e.track),
            "department": e.department,
            "last_login_at": e.last_login_at.isoformat() if e.last_login_at else None,
            "first_login_at": e.first_login_at.isoformat() if e.first_login_at else None,
            "modules_completed": emp_modules_completed.get(e.employee_id, 0),
            "point_total": pts_by_emp.get(e.employee_id),
            "threshold": _threshold_for(pts_by_emp.get(e.employee_id)),
            "is_manager": e.is_manager,
            "reports": sub_reports_by_mgr.get(e.employee_id, []),
        }
        for e in sorted(team, key=lambda x: (x.last_name, x.first_name))
    ]

    total_planned = sum(v["planned_count"] for v in absence_by_emp.values())
    total_unplanned = sum(v["unplanned_count"] for v in absence_by_emp.values())

    last_updated = None
    candidates = [t for t in [last_updated_time, last_updated_reviews, last_updated_absences] if t is not None]
    if candidates:
        last_updated = max(candidates)

    return {
        "team_size": len(team),
        "last_updated": last_updated.isoformat() if last_updated else None,
        "last_updated_time": last_updated_time.isoformat() if last_updated_time else None,
        "last_updated_reviews": last_updated_reviews.isoformat() if last_updated_reviews else None,
        "last_updated_absences": last_updated_absences.isoformat() if last_updated_absences else None,
        "hours_summary": sorted(hours_by_emp.values(), key=lambda x: x["full_name"]),
        "hours_date_range": hours_date_range,
        "hours_week_count": hours_week_count,
        "upcoming_reviews": upcoming,
        "past_due_reviews": past_due,
        "team": team_list,
        "absence_summary": sorted(absence_by_emp.values(), key=lambda x: x["full_name"]),
        "absence_by_category": sorted(absence_category_by_emp.values(), key=lambda x: x["full_name"]),
        "total_planned_absences": total_planned,
        "total_unplanned_absences": total_unplanned,
        "absence_date_range": absence_date_range,
    }


def _latest_points_by_employee(team_ids: set[str], db: Session) -> dict[str, float]:
    """Return {employee_id: point_total} using the most recently imported record per employee."""
    if not team_ids:
        return {}
    subq = (
        db.query(
            AttendancePoint.employee_id,
            sa_func.max(AttendancePoint.imported_at).label("latest_at"),
        )
        .filter(AttendancePoint.employee_id.in_(team_ids))
        .group_by(AttendancePoint.employee_id)
        .subquery()
    )
    latest = (
        db.query(AttendancePoint)
        .join(
            subq,
            (AttendancePoint.employee_id == subq.c.employee_id)
            & (AttendancePoint.imported_at == subq.c.latest_at),
        )
        .all()
    )
    return {r.employee_id: r.point_total for r in latest}


def _monday_flag_ids_for_month(team_ids: set[str], month_start: str, month_end: str, db: Session) -> set[str]:
    """Return set of employee_ids that have any non-null flag_code in the given month."""
    if not team_ids:
        return set()
    rows = db.query(AttendancePoint.employee_id).filter(
        AttendancePoint.employee_id.in_(team_ids),
        AttendancePoint.point_date >= month_start,
        AttendancePoint.point_date <= month_end,
        AttendancePoint.flag_code.isnot(None),
        AttendancePoint.flag_code != "",
    ).distinct().all()
    return {r.employee_id for r in rows}


# ── Dashboard ─────────────────────────────────────────────────────────────────

@router.get("/dashboard")
def get_manager_dashboard(
    manager: dict = Depends(require_manager),
    db: Session = Depends(get_db),
    month: str = Query(default=None, description="Month to show, format YYYY-MM. Defaults to current month."),
    compare_month: str = Query(default=None, description="Optional second month for side-by-side comparison."),
    weeks: int = Query(default=None, ge=0, le=520, description="[Deprecated] weeks look-back. Use month= instead."),
    as_manager_id: str = Query(default=None, description="Admin-only: view dashboard as a specific manager."),
):
    """Aggregated manager dashboard. Supports month-based and legacy weeks-based filtering."""
    # View-as logic — admins and executives can view any manager's dashboard
    if as_manager_id and (manager.get("is_admin") or manager.get("is_executive")):
        manager_id = as_manager_id
    else:
        manager_id = manager["sub"]

    # Active team members only — terminated employees are excluded from current view
    team = (
        db.query(Employee)
        .filter(Employee.manager_employee_id == manager_id, Employee.terminated_date.is_(None))
        .all()
    )
    team_ids = {e.employee_id for e in team}
    team_by_id = {e.employee_id: e for e in team}

    if not team_ids:
        empty = {
            "team_size": 0,
            "last_updated": None,
            "last_updated_time": None,
            "last_updated_reviews": None,
            "last_updated_absences": None,
            "hours_summary": [],
            "hours_date_range": None,
            "hours_week_count": 0,
            "upcoming_reviews": [],
            "past_due_reviews": [],
            "team": [],
            "absence_summary": [],
            "absence_by_category": [],
            "total_planned_absences": 0,
            "total_unplanned_absences": 0,
            "absence_date_range": None,
        }
        if compare_month:
            return {"month": {**empty, "month": month or _current_month()},
                    "compare_month": {**empty, "month": compare_month}}
        return empty

    # Determine month range — month param takes priority; weeks= is legacy fallback
    if month:
        month_start, month_end = _month_bounds(month)
    elif weeks is not None:
        # Legacy path: convert weeks to a date range using today as anchor
        if weeks == 0:
            month_start = "1970-01-01"
            month_end = date.today().isoformat()
        else:
            cutoff = (date.today() - timedelta(weeks=weeks)).isoformat()
            month_start = cutoff
            month_end = date.today().isoformat()
        month = month or _current_month()
    else:
        month = _current_month()
        month_start, month_end = _month_bounds(month)

    pts_by_emp = _latest_points_by_employee(team_ids, db)
    monday_flag_ids = _monday_flag_ids_for_month(team_ids, month_start, month_end, db)

    snapshot = _build_snapshot(team, team_ids, team_by_id, db, month_start, month_end, pts_by_emp, monday_flag_ids)
    snapshot["month"] = month

    if compare_month:
        cmp_start, cmp_end = _month_bounds(compare_month)
        cmp_monday_ids = _monday_flag_ids_for_month(team_ids, cmp_start, cmp_end, db)
        cmp_snapshot = _build_snapshot(team, team_ids, team_by_id, db, cmp_start, cmp_end, pts_by_emp, cmp_monday_ids)
        cmp_snapshot["month"] = compare_month
        return {"month": snapshot, "compare_month": cmp_snapshot}

    return snapshot


def _current_month() -> str:
    today = date.today()
    return f"{today.year}-{today.month:02d}"


# ── Employee detail ───────────────────────────────────────────────────────────

@router.get("/employee/{employee_id}")
def get_employee_detail(
    employee_id: str,
    manager: dict = Depends(require_manager),
    db: Session = Depends(get_db),
    month: str = Query(default=None, description="Month to show, format YYYY-MM. Defaults to current month."),
):
    """Full detail for a single team member: hours, reviews, and attendance point history."""
    manager_id = manager["sub"]

    # Admins can view any employee; managers can only view their own team
    emp = db.query(Employee).filter_by(employee_id=employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found.")

    if not manager.get("is_admin"):
        team_ids = {e.employee_id for e in db.query(Employee).filter_by(manager_employee_id=manager_id).all()}
        if employee_id not in team_ids:
            raise HTTPException(status_code=403, detail="Employee is not on your team.")

    month_str = month or _current_month()
    month_start, month_end = _month_bounds(month_str)

    # Hours for the month
    time_records = db.query(TimeRecord).filter(
        TimeRecord.employee_id == employee_id,
        TimeRecord.week_start >= month_start,
        TimeRecord.week_start <= month_end,
    ).all()

    hours: dict[str, float] = {
        "regular": 0.0, "ot": 0.0,
        "vacation": 0.0, "personal": 0.0,
        "absent_w_point": 0.0, "protected": 0.0, "other": 0.0,
    }
    for r in time_records:
        hours["regular"]        += r.regular_hours
        hours["ot"]             += r.ot_hours
        hours["vacation"]       += getattr(r, "vacation_hours", 0.0) or 0.0
        hours["personal"]       += getattr(r, "personal_hours", 0.0) or 0.0
        hours["absent_w_point"] += getattr(r, "absent_w_point_hours", 0.0) or 0.0
        hours["protected"]      += getattr(r, "protected_hours", 0.0) or 0.0
        hours["other"]          += getattr(r, "other_hours", 0.0) or 0.0
    hours = {k: round(v, 1) for k, v in hours.items()}

    # Performance reviews — all incomplete (not month-scoped)
    reviews = db.query(PerformanceReview).filter(
        PerformanceReview.employee_id == employee_id,
        PerformanceReview.completed == False,  # noqa: E712
    ).all()
    today = date.today()
    upcoming_rev: list[dict] = []
    past_due_rev: list[dict] = []
    for r in reviews:
        try:
            due = date.fromisoformat(r.due_date)
        except ValueError:
            continue
        diff = (due - today).days
        entry = {"review_type": r.review_type, "due_date": r.due_date}
        if diff >= 0:
            upcoming_rev.append({**entry, "days_until": diff})
        else:
            past_due_rev.append({**entry, "days_overdue": abs(diff)})
    upcoming_rev.sort(key=lambda x: x["days_until"])
    past_due_rev.sort(key=lambda x: x["days_overdue"], reverse=True)

    # Attendance point events for the month
    point_events = db.query(AttendancePoint).filter(
        AttendancePoint.employee_id == employee_id,
        AttendancePoint.point_date >= month_start,
        AttendancePoint.point_date <= month_end,
    ).order_by(AttendancePoint.point_date).all()

    # Current point total — most recent imported record all-time
    latest_pts = _latest_points_by_employee({employee_id}, db)
    current_point_total = latest_pts.get(employee_id)

    employee_dict = {
        "employee_id": emp.employee_id,
        "full_name": f"{emp.first_name} {emp.last_name}",
        "tracks": normalize_tracks(emp.track),
        "department": emp.department,
        "last_login_at": emp.last_login_at.isoformat() if emp.last_login_at else None,
        "first_login_at": emp.first_login_at.isoformat() if emp.first_login_at else None,
        "point_total": current_point_total,
        "threshold": _threshold_for(current_point_total),
    }

    return {
        "employee": employee_dict,
        "hours": hours,
        "reviews": {"upcoming": upcoming_rev, "past_due": past_due_rev},
        "attendance_points": [
            {
                "point_date": p.point_date,
                "point": p.point,
                "reason": p.reason,
                "note": p.note,
                "flag_code": p.flag_code,
                "point_total": p.point_total,
                "imported_at": p.imported_at.isoformat() if p.imported_at else None,
            }
            for p in point_events
        ],
        "current_point_total": current_point_total,
        "threshold": _threshold_for(current_point_total),
    }
